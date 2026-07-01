from fastapi.testclient import TestClient
from pydantic import SecretStr

from alletra_onboard.adapters.persistence.sqlite import SqliteRunStore
from alletra_onboard.api.app import create_app
from alletra_onboard.application.event_bus import InMemoryEventBus
from alletra_onboard.application.intake import csv_template
from alletra_onboard.application.onboarding_service import OnboardingService
from alletra_onboard.config import Settings
from alletra_onboard.domain.models import (
    ArrayWorkItem,
    DsccSetupConfig,
    NetworkConfig,
    VerificationReport,
)


def _client(tmp_path) -> TestClient:
    store = SqliteRunStore(tmp_path / "state.db")
    store.initialize()
    service = OnboardingService(Settings(), store, InMemoryEventBus())
    return TestClient(create_app(service))


def _work_item_payload() -> dict:
    item = ArrayWorkItem(
        serial_number="SGHD45FF0Y",
        part_number="S0B84A",
        subscription_key=SecretStr("REALKEY123"),
        service_catalog_region_id="ap-northeast",
        dscc_region_code="jp1",
        cloudinit_url="https://169.254.239.27/cloudinit",
        network=NetworkConfig(
            mgmt_ipv4="10.64.154.225",
            mask="255.255.248.0",
            gateway="10.64.159.254",
            dns=["10.203.96.10"],
            ntp="ntp.example.com",
            timezone="Asia/Kolkata",
        ),
        dscc_setup=DsccSetupConfig(system_name="array01", country="India"),
    )
    payload = item.model_dump(mode="json")
    payload["subscription_key"] = "REALKEY123"  # model_dump masks; the client sends the real key
    return payload


def test_health(tmp_path):
    assert _client(tmp_path).get("/health").json() == {"status": "ok"}


def test_create_get_and_list_runs_masks_secret(tmp_path):
    client = _client(tmp_path)
    created = client.post("/runs", json={"work_item": _work_item_payload()})
    assert created.status_code == 200
    run_id = created.json()["run"]["run_id"]

    detail = client.get(f"/runs/{run_id}")
    assert detail.status_code == 200
    body = detail.json()
    assert body["run"]["serial_number"] == "SGHD45FF0Y"
    # the API must never echo the subscription key
    assert "REALKEY123" not in detail.text

    assert any(r["run_id"] == run_id for r in client.get("/runs").json()["runs"])
    events = client.get(f"/runs/{run_id}/events").json()["events"]
    assert events and events[0]["event_type"] == "run.created"


def test_run_not_found_is_404(tmp_path):
    client = _client(tmp_path)
    assert client.get("/runs/nope").status_code == 404
    assert client.post("/runs/nope/provision", json={}).status_code == 404


def test_mark_complete(tmp_path):
    client = _client(tmp_path)
    run_id = client.post("/runs", json={"work_item": _work_item_payload()}).json()["run"]["run_id"]
    done = client.post(f"/runs/{run_id}/complete")
    assert done.json()["run"]["status"] == "succeeded"


def test_prereq_firewall_and_connectivity(tmp_path, monkeypatch):
    client = _client(tmp_path)

    fw = client.get("/prereqs/firewall?region=jp1")
    assert fw.status_code == 200
    rules = fw.json()["rules"]
    fqdns = [r["fqdn"] for r in rules]
    assert "console.greenlake.hpe.com" in fqdns
    assert "jp1.data.cloud.hpe.com" in fqdns  # <instance> substituted for the region
    assert all(r["port"] == "TCP 443" for r in rules)

    txt = client.get("/prereqs/firewall.txt?region=jp1")
    assert txt.status_code == 200
    assert "attachment" in txt.headers["content-disposition"]
    assert "console.greenlake.hpe.com" in txt.text

    # Connectivity does real network I/O — stub it so the test stays hermetic.
    from alletra_onboard.application import prereqs

    async def fake_check(region="jp1", timeout=5.0):
        return [
            prereqs.ConnectivityResult("console.greenlake.hpe.com", 443, True, "reachable"),
            prereqs.ConnectivityResult("device.cloud.hpe.com", 443, False, "blocked (ConnectionRefusedError)"),
        ]

    monkeypatch.setattr(prereqs, "check_connectivity", fake_check)
    conn = client.get("/prereqs/connectivity?region=jp1").json()
    assert conn["all_reachable"] is False
    assert {c["host"] for c in conn["results"]} == {"console.greenlake.hpe.com", "device.cloud.hpe.com"}


def test_init_sheet_template_has_prerequisites_tab(tmp_path):
    import io

    from openpyxl import load_workbook

    resp = _client(tmp_path).get("/init-sheet/template")
    assert resp.status_code == 200
    wb = load_workbook(io.BytesIO(resp.content))
    assert wb.sheetnames[0] == "Initialisation"  # the fillable tab stays first (the parser reads it)
    assert "Prerequisites" in wb.sheetnames


def test_verify_endpoint_accepts_credentials_and_validates(tmp_path):
    # Inject a stub verify_fn so the endpoint never attempts a real SSH connection.
    store = SqliteRunStore(tmp_path / "state.db")
    store.initialize()
    service = OnboardingService(
        Settings(), store, InMemoryEventBus(),
        verify_fn=lambda item, username, password: VerificationReport(reachable=True),
    )
    client = TestClient(create_app(service))
    run_id = client.post("/runs", json={"work_item": _work_item_payload()}).json()["run"]["run_id"]
    client.post(f"/runs/{run_id}/complete")

    ok = client.post(f"/runs/{run_id}/verify", json={"username": "3paradm", "password": "pw"})
    assert ok.status_code == 200

    # the password is required — a read-only check still has to authenticate
    assert client.post(f"/runs/{run_id}/verify", json={"username": "3paradm"}).status_code == 422
    # unknown run -> 404
    assert client.post("/runs/nope/verify", json={"username": "u", "password": "p"}).status_code == 404


def test_template_and_parse_round_trip(tmp_path):
    client = _client(tmp_path)
    template = client.get("/work-items/template")
    assert template.status_code == 200
    assert template.text == csv_template()

    parsed = client.post("/work-items/parse", json={"csv_text": template.text})
    assert parsed.status_code == 200
    items = parsed.json()["work_items"]
    assert len(items) == 1 and items[0]["serial_number"] == "SGHD00EXAMPLE"
    # parse echoes back what the operator uploaded (for the editable form) — NOT masked...
    assert items[0]["subscription_key"] == "EXAMPLEKEY1234567890"
    # ...except passwords, which are never echoed.
    assert "password" not in items[0]["dscc_setup"]

    bad = client.post("/work-items/parse", json={"csv_text": "not,a,real\nheader,row,x"})
    assert bad.status_code == 422


def test_init_sheet_template_and_upload(tmp_path, monkeypatch):
    import base64
    import io

    from openpyxl import load_workbook

    from alletra_onboard.application.configuring import read_env
    from alletra_onboard.application.init_sheet import SECTIONS

    monkeypatch.chdir(tmp_path)  # the API writes .env in the working directory
    client = _client(tmp_path)

    template = client.get("/init-sheet/template")
    assert template.status_code == 200
    assert "spreadsheetml" in template.headers["content-type"]

    values = {
        "gl_client_id": "client-123", "gl_client_secret": "secret-xyz",
        "gl_token_url": "https://global.api.greenlake.hpe.com/authorization/v2/oauth2/t/token",
        "serial_number": "SGHD45FF0Y", "part_number": "S0B84A", "subscription_key": "YHHDKEY123",
        "service_catalog_region_id": "ap-northeast", "dscc_region_code": "jp1",
        "mgmt_ipv4": "10.64.154.225", "mask": "255.255.248.0", "gateway": "10.64.159.254",
        "dns1": "10.203.96.10", "ntp": "ntp1.example.net", "timezone": "Asia/Kolkata",
        "contact_first_name": "Jane", "contact_last_name": "Doe", "contact_language": "English",
        "contact_company": "HPE", "contact_phone": "8000000000", "contact_email": "jane@example.com",
        "dscc_system_name": "MPB10K-TEST", "dscc_country": "India",
        "secret_name": "b10000-admin", "secret_username": "3paradm",
    }
    label_to_key = {label: key for _, fields in SECTIONS for key, label, _, _ in fields}
    wb = load_workbook(io.BytesIO(template.content))
    for row in wb.active.iter_rows(min_row=2):
        if row[0].value is None:
            continue
        key = label_to_key.get(str(row[0].value).strip().removesuffix("*").strip())
        if key in values:
            row[1].value = values[key]
    buffer = io.BytesIO()
    wb.save(buffer)
    b64 = base64.b64encode(buffer.getvalue()).decode()

    resp = client.post("/init-sheet/upload", json={"content_b64": b64})
    assert resp.status_code == 200
    body = resp.json()
    assert body["run"]["serial_number"] == "SGHD45FF0Y"
    assert body["credentials_saved"] is True
    # the admin password is not in the sheet at all, and is never echoed to the UI
    assert "password" not in body["work_item"]["dscc_setup"]
    # GreenLake credentials from the sheet were written to .env
    env = read_env(tmp_path / ".env")
    assert env["GL_CLIENT_ID"] == "client-123"
    assert env["GL_CLIENT_SECRET"] == "secret-xyz"

    # a non-xlsx upload is a clean 422
    bad = client.post("/init-sheet/upload", json={"content_b64": base64.b64encode(b"nope").decode()})
    assert bad.status_code == 422


def _fill_template(template_bytes: bytes, values: dict[str, str]) -> str:
    """Fill the Value column of the Initialisation tab for the given keys -> base64 xlsx."""
    import base64
    import io

    from openpyxl import load_workbook

    from alletra_onboard.application.init_sheet import SECTIONS

    label_to_key = {label: key for _, fields in SECTIONS for key, label, _, _ in fields}
    wb = load_workbook(io.BytesIO(template_bytes))
    for row in wb.active.iter_rows(min_row=2):
        if row[0].value is None:
            continue
        key = label_to_key.get(str(row[0].value).strip().removesuffix("*").strip())
        if key in values:
            row[1].value = values[key]
    buffer = io.BytesIO()
    wb.save(buffer)
    return base64.b64encode(buffer.getvalue()).decode()


def test_verify_only_upload_lands_on_verify_without_greenlake_creds(tmp_path, monkeypatch):
    # Decoupling: a sparse sheet (serial + array IP only) + mode=VERIFY_ONLY creates a run that
    # resumes on the verify step and does NOT write GreenLake creds to .env.
    from alletra_onboard.application.configuring import read_env

    monkeypatch.chdir(tmp_path)
    client = _client(tmp_path)
    template = client.get("/init-sheet/template").content
    b64 = _fill_template(template, {"serial_number": "SGHD45FF0Y", "mgmt_ipv4": "10.64.154.225"})

    resp = client.post("/init-sheet/upload", json={"content_b64": b64, "mode": "VERIFY_ONLY"})
    assert resp.status_code == 200, resp.text
    run = resp.json()["run"]
    assert run["mode"] == "VERIFY_ONLY"
    assert run["current_phase"] == "CONFIG_VERIFY"  # resumes straight to the SSH check

    # no GreenLake creds in the sheet -> nothing written to .env
    env_path = tmp_path / ".env"
    assert not env_path.exists() or "GL_CLIENT_ID" not in read_env(env_path)


def test_full_mode_upload_still_requires_every_field(tmp_path, monkeypatch):
    # Default/full mode keeps the strict validation: a sparse sheet is a clean 422.
    monkeypatch.chdir(tmp_path)
    client = _client(tmp_path)
    template = client.get("/init-sheet/template").content
    b64 = _fill_template(template, {"serial_number": "SGHD45FF0Y", "mgmt_ipv4": "10.64.154.225"})
    resp = client.post("/init-sheet/upload", json={"content_b64": b64, "mode": "FULL_ONBOARDING"})
    assert resp.status_code == 422
    assert "API Client" in resp.text  # missing GreenLake creds reported


def test_provisioning_writes_frozen_by_default(tmp_path, monkeypatch):
    monkeypatch.delenv("PROVISIONING_WRITES_ENABLED", raising=False)
    client = _client(tmp_path)
    caps = client.get("/provisioning/capabilities").json()
    assert caps["writes_enabled"] is False

    run_id = client.post("/runs", json={"work_item": _work_item_payload()}).json()["run"]["run_id"]
    # apply (a write) is refused with 403 before it even looks for a plan/intent
    resp = client.post(f"/runs/{run_id}/storage/apply")
    assert resp.status_code == 403
    assert "frozen" in resp.text.lower()


def test_config_roundtrip_masks_secret(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)  # the API writes .env in the working directory
    client = _client(tmp_path)
    saved = client.post(
        "/config",
        json={"gl_client_id": "abc", "gl_client_secret": "supersecret", "gl_token_url": "https://t"},
    )
    assert saved.status_code == 200
    body = saved.json()
    assert body["configured"] is True
    assert body["values"]["GL_CLIENT_ID"] == "abc"
    assert body["values"]["GL_CLIENT_SECRET"] == "****"
    assert "supersecret" not in saved.text
