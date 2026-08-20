import io

import pytest
from openpyxl import load_workbook

from alletra_onboard.application.platform.init_sheet import (
    HOSTSET_COLUMNS,
    HOSTSETS_SHEET_NAME,
    PROVISIONING_SECTIONS,
    PROVISIONING_SHEET_NAME,
    SECTIONS,
    VOLUME_COLUMNS,
    VOLUMES_SHEET_NAME,
    build_template_bytes,
    parse_workbook_bytes,
)
from alletra_onboard.domain.models import RunMode

# The provisioning input now spans three tabs: Provisioning (targets + defaults, key-value), Volumes
# (row-table), Host sets (row-table). Tests supply a dict {"targets", "volumes", "hostsets"}.
_PROV_TARGETS = {
    "prov_array_host": "10.64.122.140", "prov_array_user": "3paradm", "prov_array_password": "pw",
    "prov_vcenter_host": "vc.example.net", "prov_vcenter_user": "administrator@vsphere.local", "prov_vcenter_password": "vpw",
    "prov_sw1_host": "10.0.0.1", "prov_sw1_user": "admin", "prov_sw1_password": "s1",
    "prov_sw2_host": "10.0.0.2", "prov_sw2_user": "admin", "prov_sw2_password": "s2",
}
_PROV_VOLUMES = [
    {"name": "CRV_LZ_Prod01", "size_gib": "1024"},
    {"name": "CRV_LZ_Prod02", "size_gib": "1024"},
    {"name": "CRV_LZ_Prod03", "size_gib": "1024"},
]
_PROV_HOSTSETS = [{"name": "CRVLZ_Hostset"}]
_PROV_COMPLETE = {"targets": _PROV_TARGETS, "volumes": _PROV_VOLUMES, "hostsets": _PROV_HOSTSETS}


def _norm(s: object) -> str:
    return str(s).strip().removesuffix("*").strip()


def _fill_row_table(ws, columns, records: list[dict[str, str]]) -> None:
    """Locate the header row (by matching column labels) and write each record into the rows beneath."""
    header_to_key = {_norm(label): key for key, label, _ in columns}
    header_row, col_of_key = None, {}
    for r in ws.iter_rows():
        found = {header_to_key[_norm(c.value)]: c.column for c in r if c.value is not None and _norm(c.value) in header_to_key}
        if found:
            header_row, col_of_key = r[0].row, found
            break
    for i, rec in enumerate(records):
        for key, val in rec.items():
            ws.cell(row=header_row + 1 + i, column=col_of_key[key], value=val)


def _fill_tabs(init_values: dict[str, str], prov: dict | None = None) -> bytes:
    """Fill the Initialisation tab and (optionally) the Provisioning + Volumes + Host sets tabs."""
    init_l2k = {label: key for _, fields in SECTIONS for key, label, _, _ in fields}
    prov_l2k = {label: key for _, fields in PROVISIONING_SECTIONS for key, label, _, _ in fields}
    wb = load_workbook(io.BytesIO(build_template_bytes()))

    def fill_kv(ws, label_to_key, vals):
        for row in ws.iter_rows(min_row=2):
            if row[0].value is None:
                continue
            key = label_to_key.get(_norm(row[0].value))
            if key and key in vals:
                row[1].value = vals[key]

    fill_kv(wb["Initialisation"], init_l2k, init_values)
    if prov is not None:
        fill_kv(wb[PROVISIONING_SHEET_NAME], prov_l2k, prov.get("targets", {}))
        _fill_row_table(wb[VOLUMES_SHEET_NAME], VOLUME_COLUMNS, prov.get("volumes", []))
        _fill_row_table(wb[HOSTSETS_SHEET_NAME], HOSTSET_COLUMNS, prov.get("hostsets", []))
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _fill(values: dict[str, str]) -> bytes:
    """Take the blank template and fill the Value column for the given field keys."""
    label_to_key = {label: key for _, fields in SECTIONS for key, label, _, _ in fields}
    wb = load_workbook(io.BytesIO(build_template_bytes()))
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        label = row[0].value
        if label is None:
            continue
        key = label_to_key.get(str(label).strip().removesuffix("*").strip())
        if key and key in values:
            row[1].value = values[key]
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


_COMPLETE = {
    "gl_client_id": "client-123",
    "gl_client_secret": "secret-xyz",
    "gl_token_url": "https://global.api.greenlake.hpe.com/authorization/v2/oauth2/tenant/token",
    "serial_number": "SGHD45FF0Y",
    "part_number": "S0B84A",
    "subscription_key": "YHHDKEY1234567890",
    "service_catalog_region_id": "ap-northeast",
    "dscc_region_code": "jp1",
    "mgmt_ipv4": "10.64.154.225",
    "mask": "255.255.248.0",
    "gateway": "10.64.159.254",
    "dns1": "10.203.96.10",
    "dns2": "10.203.96.9",
    "ntp": "ntp1.example.net",
    "timezone": "Asia/Kolkata",
    "proxy_host": "proxy.example.net",
    "proxy_port": "8080",
    "contact_first_name": "Jane",
    "contact_last_name": "Doe",
    "contact_language": "English",
    "contact_company": "HPE",
    "contact_phone": "8000000000",
    "contact_email": "jane.doe@example.com",
    "dscc_system_name": "MPB10K-TEST",
    "dscc_country": "India",
    "secret_name": "b10000-admin",
    "secret_username": "3paradm",
}


def test_template_is_a_valid_xlsx_with_every_field():
    wb = load_workbook(io.BytesIO(build_template_bytes()))
    labels = {str(row[0].value).strip().removesuffix("*").strip() for row in wb.active.iter_rows(min_row=2) if row[0].value}
    for _, fields in SECTIONS:
        for _, label, _, _ in fields:
            assert label in labels


def test_round_trip_parses_creds_and_work_item():
    parsed = parse_workbook_bytes(_fill(_COMPLETE))
    assert parsed.gl_client_id == "client-123"
    assert parsed.gl_client_secret == "secret-xyz"
    assert parsed.gl_token_url.endswith("/token")

    item = parsed.work_item
    assert item.serial_number == "SGHD45FF0Y"
    assert item.part_number == "S0B84A"
    assert item.subscription_key.get_secret_value() == "YHHDKEY1234567890"
    assert item.network.mgmt_ipv4 == "10.64.154.225"
    assert item.network.dns == ["10.203.96.10", "10.203.96.9"]
    assert item.network.proxy_port == 8080
    # the admin password is NOT in the sheet — the operator types it in the DSCC wizard
    assert item.dscc_setup.password is None
    assert item.dscc_setup.system_name == "MPB10K-TEST"


def test_missing_required_field_is_reported():
    incomplete = dict(_COMPLETE)
    del incomplete["gl_client_secret"]
    del incomplete["mgmt_ipv4"]
    with pytest.raises(ValueError) as exc:
        parse_workbook_bytes(_fill(incomplete))
    assert "API Client Secret" in str(exc.value)
    assert "IP address" in str(exc.value)


def test_verify_only_sheet_needs_just_serial_and_ip():
    # A sparse sheet (no GreenLake creds, no DSCC contact) parses fine for a verify-only run.
    sparse = {"serial_number": "SGHD45FF0Y", "mgmt_ipv4": "10.64.154.225"}
    parsed = parse_workbook_bytes(_fill(sparse), mode=RunMode.VERIFY_ONLY)
    assert parsed.work_item.serial_number == "SGHD45FF0Y"
    assert parsed.work_item.network.mgmt_ipv4 == "10.64.154.225"
    assert parsed.gl_client_id == ""  # not required, not present


def test_verify_only_still_requires_the_array_ip():
    with pytest.raises(ValueError) as exc:
        parse_workbook_bytes(_fill({"serial_number": "SGHD45FF0Y"}), mode=RunMode.VERIFY_ONLY)
    assert "IP address" in str(exc.value)


def test_provision_only_parses_the_provisioning_tabs_not_greenlake():
    parsed = parse_workbook_bytes(
        _fill_tabs({"serial_number": "SGHD45FF0Y", "mgmt_ipv4": "10.64.122.140"}, _PROV_COMPLETE),
        mode=RunMode.PROVISION_ONLY,
    )
    assert parsed.gl_client_id == ""  # GreenLake creds not required for provisioning
    intent = parsed.provisioning_intent
    assert intent is not None
    assert [hs.name for hs in intent.host_sets] == ["CRVLZ_Hostset"]
    assert intent.host_sets[0].members == []  # blank Members => all discovered hosts
    assert intent.array.host == "10.64.122.140"
    assert intent.array.password.get_secret_value() == "pw"
    assert intent.switch_f1.host == "10.0.0.1" and intent.switch_f2.host == "10.0.0.2"
    assert [v.name for v in intent.volumes] == ["CRV_LZ_Prod01", "CRV_LZ_Prod02", "CRV_LZ_Prod03"]
    assert intent.volumes[0].provisioning_type == "tpvv" and intent.volumes[0].cpg == "SSD_r6"  # defaults
    assert intent.exports == []  # exports are composed later in the UI, never in the sheet


def test_provisioning_row_tables_are_plural_and_heterogeneous():
    prov = {
        "targets": _PROV_TARGETS,
        "volumes": [
            {"name": "boot", "size_gib": "64", "provisioning_type": "tpvv", "cpg": "SSD_r6", "vvset": "cluster_vvset"},
            {"name": "data", "size_gib": "4096", "provisioning_type": "reduce", "cpg": "NL_r6", "vvset": "cluster_vvset"},
            {"name": "scratch", "size_gib": "512"},  # Type/CPG blank -> defaults
        ],
        "hostsets": [
            {"name": "prod_cluster", "members": "esx1, esx2 , esx3"},
            {"name": "dr_cluster"},  # blank members -> all discovered
        ],
    }
    intent = parse_workbook_bytes(
        _fill_tabs({"serial_number": "SGHD45FF0Y", "mgmt_ipv4": "10.64.122.140"}, prov),
        mode=RunMode.PROVISION_ONLY,
    ).provisioning_intent
    assert [(v.name, v.size_gib, v.provisioning_type, v.cpg, v.vvset) for v in intent.volumes] == [
        ("boot", 64, "tpvv", "SSD_r6", "cluster_vvset"),
        ("data", 4096, "reduce", "NL_r6", "cluster_vvset"),
        ("scratch", 512, "tpvv", "SSD_r6", None),  # defaults applied
    ]
    assert intent.host_sets[0].members == ["esx1", "esx2", "esx3"]  # split + trimmed
    assert intent.host_sets[1].members == []


def test_a_workbook_naming_two_different_arrays_is_refused():
    """A run covers ONE array. The management IP is stated on the Initialisation tab and again on
    the Provisioning tab (the second exists so a PROVISION_ONLY run has an address at all), and
    nothing used to tie them together — so a workbook could name two arrays and the run half-worked:
    verify checked one box while the as-built read the other and timed out. MEASURED LIVE:
    init 10.132.30.121 (rack13arcus, freshly onboarded) vs provisioning 10.64.186.90.
    """
    prov = {"targets": dict(_PROV_TARGETS), "volumes": _PROV_VOLUMES, "hostsets": _PROV_HOSTSETS}
    prov["targets"]["prov_array_host"] = "10.64.186.90"
    with pytest.raises(ValueError) as exc:
        parse_workbook_bytes(
            _fill_tabs({"serial_number": "CZ2D2K014S", "mgmt_ipv4": "10.132.30.121"}, prov),
            mode=RunMode.PROVISION_ONLY,
        )
    message = str(exc.value)
    assert "10.132.30.121" in message and "10.64.186.90" in message   # BOTH values, so it is fixable
    assert "one array" in message.lower()


def test_the_same_array_in_both_tabs_is_accepted():
    prov = {"targets": dict(_PROV_TARGETS), "volumes": _PROV_VOLUMES, "hostsets": _PROV_HOSTSETS}
    prov["targets"]["prov_array_host"] = "10.132.30.121"
    parsed = parse_workbook_bytes(
        _fill_tabs({"serial_number": "CZ2D2K014S", "mgmt_ipv4": "10.132.30.121"}, prov),
        mode=RunMode.PROVISION_ONLY,
    )
    assert parsed.work_item.network.mgmt_ipv4 == "10.132.30.121"
    assert parsed.provisioning_intent.array.host == "10.132.30.121"


def test_only_the_provisioning_address_is_not_a_conflict():
    """Every mode that runs verify or cloudinit requires the Initialisation IP, so the two normally
    coexist. A run of provisioning steps ALONE has no Network section, and the Provisioning tab's
    address is then the only one stated — which is not a disagreement."""
    prov = {"targets": _PROV_TARGETS, "volumes": _PROV_VOLUMES, "hostsets": _PROV_HOSTSETS}
    parsed = parse_workbook_bytes(
        _fill_tabs({"serial_number": "SGHD45FF0Y"}, prov),
        mode=RunMode.CUSTOM, selected_steps=["discover", "zoning", "provision"],
    )
    assert parsed.work_item.network.mgmt_ipv4 == ""
    assert parsed.provisioning_intent.array.host == "10.64.122.140"


def test_provision_only_reports_missing_target_fields():
    incomplete = {"targets": dict(_PROV_TARGETS), "volumes": _PROV_VOLUMES, "hostsets": _PROV_HOSTSETS}
    del incomplete["targets"]["prov_array_password"]
    with pytest.raises(ValueError) as exc:
        parse_workbook_bytes(
            _fill_tabs({"serial_number": "SGHD45FF0Y", "mgmt_ipv4": "10.64.122.140"}, incomplete),
            mode=RunMode.PROVISION_ONLY,
        )
    assert "Provisioning tab" in str(exc.value)
    assert "Array admin password" in str(exc.value)


def test_provision_only_requires_at_least_one_volume():
    no_vols = {"targets": _PROV_TARGETS, "volumes": [], "hostsets": _PROV_HOSTSETS}
    with pytest.raises(ValueError, match="Volumes tab"):
        parse_workbook_bytes(
            _fill_tabs({"serial_number": "SGHD45FF0Y", "mgmt_ipv4": "10.64.122.140"}, no_vols),
            mode=RunMode.PROVISION_ONLY,
        )


def test_full_mode_ignores_the_provisioning_tabs():
    # No provisioning step selected -> the Provisioning tabs are not parsed at all.
    parsed = parse_workbook_bytes(_fill_tabs(_COMPLETE))  # default FULL_ONBOARDING, no prov values
    assert parsed.provisioning_intent is None
