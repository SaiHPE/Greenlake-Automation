"""The operator's single input: ``Initialisation_sheet.xlsx``.

It mirrors HPE's official "Initialization Worksheet" (so it's familiar to the engineer) and adds
a GreenLake-API section, so one filled workbook carries BOTH the workspace API credentials and
all per-array parameters. One workbook = one array.

``build_template_bytes()`` produces the blank template to hand the customer; ``parse_workbook_bytes()``
reads a filled one back into (GreenLake creds, ArrayWorkItem). The operator fills the **Value**
column only — the **Field** column is the stable key the parser matches on, so it must not be renamed.
"""

from __future__ import annotations

import io
from dataclasses import dataclass

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from alletra_onboard.application.platform import prereqs
from alletra_onboard.domain.models import ArrayWorkItem, DsccSetupConfig, NetworkConfig, RunMode
from alletra_onboard.domain.shared import EndpointCreds
from alletra_onboard.domain.provisioning import HostSetRequest, ProvisioningIntent, VolumeRequest
from alletra_onboard.domain.workflow import enabled_steps

SHEET_NAME = "Initialisation"
# The link-local cloudinit URL changes every boot, so it is NOT in the sheet — the operator pastes
# the fresh one in the Cloud Connectivity step. This placeholder keeps the work item valid until then.
DEFAULT_CLOUDINIT_URL = "https://169.254.0.0/cloudinit"
_HPE_GREEN = "01A982"

# section title -> list of (key, label, required, note). `key` is the stable identifier; `label`
# is what appears in the Field column (the parser strips a trailing " *" required marker).
SECTIONS: list[tuple[str, list[tuple[str, str, bool, str]]]] = [
    ("HPE GreenLake API access", [
        ("gl_client_id", "API Client ID", True, "GreenLake > Manage > API clients (create one)"),
        ("gl_client_secret", "API Client Secret", True, "shown once when you create the API client"),
        ("gl_token_url", "API Token URL", True, "https://global.api.greenlake.hpe.com/authorization/v2/oauth2/<tenant>/token"),
    ]),
    ("Array (cloud enablement)", [
        ("serial_number", "Storage system serial number", True, "e.g. SGHD45FF0Y"),
        ("part_number", "Product number (SKU)", True, "product SKU e.g. S0B84A — NOT the box FRU"),
        ("subscription_key", "Subscription key", True, "GreenLake activation key from the ESD receipt email"),
        ("service_catalog_region_id", "Data Services region", True, "e.g. ap-northeast"),
        ("dscc_region_code", "DSCC region code", True, "e.g. jp1"),
    ]),
    ("Network", [
        ("mgmt_ipv4", "IP address", True, "array management IPv4 address"),
        ("mask", "Netmask", True, "e.g. 255.255.248.0"),
        ("gateway", "Gateway", True, ""),
        ("dns1", "DNS server 1", True, ""),
        ("dns2", "DNS server 2", False, "optional"),
        ("dns3", "DNS server 3", False, "optional"),
        ("ntp", "Time (NTP) server", True, "must keep array time within 2 minutes of correct"),
        ("timezone", "Timezone", True, "e.g. Asia/Kolkata"),
        ("proxy_host", "HTTP proxy server", False, "leave blank for no proxy"),
        ("proxy_port", "HTTP proxy port", False, "e.g. 8080"),
    ]),
    ("DSCC Setup — Support Contact", [
        ("contact_first_name", "First name", True, ""),
        ("contact_last_name", "Last name", True, ""),
        ("contact_language", "Preferred language", True, "e.g. English"),
        ("contact_company", "Company", True, ""),
        ("contact_phone", "Phone number", True, ""),
        ("contact_email", "Contact email", True, ""),
    ]),
    ("DSCC Setup — System", [
        ("dscc_system_name", "System name", True, "the name DSCC will show for this array"),
        ("dscc_country", "System country location", True, "e.g. India"),
    ]),
    ("DSCC Setup — System Credentials", [
        ("secret_name", "Credential name", True, "e.g. b10000-admin"),
        ("secret_username", "Administrator name", True, "e.g. 3paradm — you enter the PASSWORD in the DSCC wizard, not here"),
    ]),
    ("As-built document", [
        ("customer_name", "Customer name", False, "goes on the as-built cover page (the array doesn't store it)"),
        ("site", "Site / location", False, "deployment site/location for the as-built (optional)"),
    ]),
]

_LABEL_TO_KEY = {label: key for _, fields in SECTIONS for key, label, _, _ in fields}
_KEY_TO_LABEL = {key: label for _, fields in SECTIONS for key, label, _, _ in fields}

# Decoupling: which sheet fields each step actually needs. A run's mode -> enabled steps -> the
# union of these is what we validate, so a verify-only / provision-only sheet isn't forced to carry
# GreenLake creds or DSCC contact details. serial_number is always required (it names the run).
_ALWAYS_REQUIRED: tuple[str, ...] = ("serial_number",)
_STEP_REQUIRES: dict[str, tuple[str, ...]] = {
    "greenlake": (
        "gl_client_id", "gl_client_secret", "gl_token_url",
        "part_number", "subscription_key", "service_catalog_region_id",
    ),
    "cloudinit": ("mgmt_ipv4", "mask", "gateway", "dns1", "ntp", "timezone"),
    "dscc": (
        "dscc_region_code",
        "contact_first_name", "contact_last_name", "contact_language",
        "contact_company", "contact_phone", "contact_email",
        "dscc_system_name", "dscc_country", "secret_name", "secret_username",
    ),
    # Provisioning steps validate against the separate 'Provisioning' tab (parsed below), so they
    # add no Initialisation-tab requirements beyond the always-required serial. Verify reuses the
    # onboarded network values, so it needs the array IP from the Initialisation tab.
    "discover": (),
    "zoning": (),
    "provision": (),
    "verify": ("mgmt_ipv4",),
}

# The provisioning input (Phase 2 / ADR 0010 Stage 2) spans THREE fillable tabs, parsed only when a
# provisioning step is selected: 'Provisioning' carries the device targets (with passwords —
# customer-supplied) + defaults; 'Volumes' and 'Host sets' are ROW-TABLES (one object per row) so a
# run can create a plural/heterogeneous set of volumes and one or more host sets. The exports
# (which source is presented to which target, at which LUN) are composed later in the operator's
# dropdown builder over discovered + to-be-created objects — they are not in the sheet.
PROVISIONING_SHEET_NAME = "Provisioning"
VOLUMES_SHEET_NAME = "Volumes"
HOSTSETS_SHEET_NAME = "Host sets"

PROVISIONING_SECTIONS: list[tuple[str, list[tuple[str, str, bool, str]]]] = [
    ("Targets — array", [
        ("prov_array_host", "Array management IP", True, "the B10000 mgmt IP (WSAPI + SSH)"),
        ("prov_array_user", "Array admin username", True, "e.g. 3paradm"),
        ("prov_array_password", "Array admin password", True, "used for WSAPI + SSH this run"),
    ]),
    ("Targets — vCenter", [
        ("prov_vcenter_host", "vCenter host or IP", True, "for read-only ESXi HBA discovery"),
        ("prov_vcenter_user", "vCenter username", True, "e.g. administrator@vsphere.local"),
        ("prov_vcenter_password", "vCenter password", True, ""),
    ]),
    ("Targets — SAN switches (dual fabric) — OPTIONAL, only for zoning remediation", [
        ("prov_sw1_host", "Switch 1 — odd / F1 — IP", False, "only needed to CREATE zones; verify is array-side"),
        ("prov_sw1_user", "Switch 1 username", False, "e.g. admin"),
        ("prov_sw1_password", "Switch 1 password", False, ""),
        ("prov_sw2_host", "Switch 2 — even / F2 — IP", False, "only needed to CREATE zones; verify is array-side"),
        ("prov_sw2_user", "Switch 2 username", False, "e.g. admin"),
        ("prov_sw2_password", "Switch 2 password", False, ""),
    ]),
    ("Defaults — used when a Volumes row leaves CPG / Type blank (optional)", [
        ("prov_default_cpg", "Default CPG", False, "default SSD_r6"),
        ("prov_default_type", "Default provisioning type", False, "tpvv (thin) or reduce (dedup+compress); default tpvv"),
    ]),
]

# Row-table columns for the Volumes / Host sets tabs: (key, header, required).
VOLUME_COLUMNS: list[tuple[str, str, bool]] = [
    ("name", "Volume name", True),
    ("size_gib", "Size (GiB)", True),
    ("provisioning_type", "Type (tpvv / reduce)", False),
    ("cpg", "CPG", False),
    ("vvset", "VV-set (optional)", False),
]
HOSTSET_COLUMNS: list[tuple[str, str, bool]] = [
    ("name", "Host set name", True),
    ("members", "Members — comma-separated host names; blank = all discovered", False),
]

_PROV_LABEL_TO_KEY = {label: key for _, fields in PROVISIONING_SECTIONS for key, label, _, _ in fields}
_PROV_KEY_TO_LABEL = {key: label for _, fields in PROVISIONING_SECTIONS for key, label, _, _ in fields}
_PROV_REQUIRED = [key for _, fields in PROVISIONING_SECTIONS for key, _, required, _ in fields if required]


def required_keys_for(mode: RunMode, selected_steps: list[str] | None = None) -> set[str]:
    """The sheet field keys a run of this mode must have filled."""
    keys = set(_ALWAYS_REQUIRED)
    for step in enabled_steps(mode, selected_steps):
        keys.update(_STEP_REQUIRES.get(step.key, ()))
    return keys


def all_required_keys() -> set[str]:
    """Every main-tab field a COMPLETE sheet must have — the union across all steps, mode-independent.
    RunMode.BOTH enables the whole step registry, so its required set is the full superset. The upload
    validates against this (see ADR 0005 revision): the sheet is filled once, in full, BEFORE a mode
    is chosen, so an incomplete sheet is rejected regardless of the intent picked later."""
    return required_keys_for(RunMode.BOTH)


@dataclass
class ParsedInitSheet:
    work_item: ArrayWorkItem
    gl_client_id: str
    gl_client_secret: str
    gl_token_url: str
    provisioning_intent: ProvisioningIntent | None = None  # set when a provisioning step is selected
    customer_name: str = ""  # for the as-built cover page (operator-supplied; not on the array)
    site: str = ""


def _provisioning_selected(mode: RunMode, selected_steps: list[str] | None) -> bool:
    return any(step.kind == "provision" for step in enabled_steps(mode, selected_steps))


def _normalize_label(text: str) -> str:
    return text.strip().removesuffix("*").strip()


def _write_fillable_tab(ws, sections: list[tuple[str, list[tuple[str, str, bool, str]]]]) -> None:
    """Write a Field | Value | Notes tab grouped by section header (the operator fills Value)."""
    ws.append(["Field", "Value", "Notes / example"])
    for cell in ws[1]:
        cell.font = Font(bold=True)

    header_fill = PatternFill("solid", fgColor=_HPE_GREEN)
    for section, fields in sections:
        row = ws.max_row + 1
        for col in (1, 2, 3):
            cell = ws.cell(row=row, column=col, value=section if col == 1 else None)
            cell.fill = header_fill
            cell.font = Font(bold=True, color="FFFFFF")
        for _key, label, required, note in fields:
            display = f"{label} *" if required else label
            ws.append([display, "", note])
            ws.cell(row=ws.max_row, column=3).font = Font(italic=True, color="808080")

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 64
    for row in ws.iter_rows():
        row[1].alignment = Alignment(wrap_text=False)
    ws.freeze_panes = "A2"


def build_template_bytes(*, init_only: bool = False) -> bytes:
    """A blank Initialisation_sheet.xlsx: Initialisation (fillable) + Prerequisites, plus the
    Provisioning tab unless ``init_only`` (the Initialization accelerator ships an onboarding-only
    sheet — no array/vCenter/switch credentials)."""
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    _write_fillable_tab(ws, SECTIONS)

    if not init_only:
        _add_provisioning_sheet(wb)
    _add_prereq_sheet(wb)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _write_table_tab(ws, columns: list[tuple[str, str, bool]], *, blank_rows: int, intro: str) -> None:
    """Write a ROW-TABLE tab: an intro line, a bold header row (one column per field), then blank rows
    the operator fills — one object per row (ADR 0010 Stage 2)."""
    ws.append([intro])
    ws.cell(row=1, column=1).font = Font(italic=True, color="808080")

    header_fill = PatternFill("solid", fgColor=_HPE_GREEN)
    header = [f"{label} *" if required else label for _key, label, required in columns]
    ws.append(header)
    header_row = ws.max_row
    for cell in ws[header_row]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")

    # Materialise blank rows as a bordered, fillable grid (append([None,...]) creates no cells, so the
    # operator would otherwise see just a header with nowhere obvious to type).
    thin = Side(style="thin", color="D0D0D0")
    grid = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(header_row + 1, header_row + 1 + blank_rows):
        for c in range(1, len(columns) + 1):
            ws.cell(row=r, column=c).border = grid

    for i, (_key, label, _required) in enumerate(columns, start=1):
        ws.column_dimensions[ws.cell(row=header_row, column=i).column_letter].width = max(20, len(label) + 4)
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate


def _add_provisioning_sheet(wb: Workbook) -> None:
    """The fillable provisioning input — parsed only when a provisioning step is selected: the
    'Provisioning' tab (targets + defaults) plus the 'Volumes' and 'Host sets' row-tables."""
    _write_fillable_tab(wb.create_sheet(PROVISIONING_SHEET_NAME), PROVISIONING_SECTIONS)
    _write_table_tab(
        wb.create_sheet(VOLUMES_SHEET_NAME), VOLUME_COLUMNS, blank_rows=15,
        intro="Volumes to create — one per row. Type/CPG blank = the Provisioning-tab defaults.",
    )
    _write_table_tab(
        wb.create_sheet(HOSTSETS_SHEET_NAME), HOSTSET_COLUMNS, blank_rows=8,
        intro="Host sets to create — one per row. Leave Members blank to include all discovered hosts.",
    )


def _add_prereq_sheet(wb: Workbook) -> None:
    """A reference 'Prerequisites' tab: firewall rules to open + a short readiness checklist.

    The parser ignores this tab (it only reads the 'Initialisation' sheet), so it is documentation
    only. The firewall list is the same single source used by the app and the downloadable .txt.
    """
    ws = wb.create_sheet("Prerequisites")
    header_fill = PatternFill("solid", fgColor=_HPE_GREEN)

    def band(text: str) -> None:
        ws.append([text, None, None])
        for col in (1, 2, 3):
            cell = ws.cell(row=ws.max_row, column=col)
            cell.fill = header_fill
            cell.font = Font(bold=True, color="FFFFFF")

    def note(text: str) -> None:
        ws.append([text, None, None])
        ws.cell(row=ws.max_row, column=1).font = Font(italic=True, color="808080")

    ws.append(["Prerequisites — complete before onboarding"])
    ws.cell(row=1, column=1).font = Font(bold=True, size=13)

    band("Firewall — open these OUTBOUND rules (send to your network team)")
    ws.append(["Destination (FQDN)", "Port", "Purpose"])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
    for fqdn, port, _initiator, purpose in prereqs.rules_for("<instance>"):
        ws.append([fqdn, port, purpose])
    note("<instance> = your DSCC region (e.g. jp1).  Also required: mDNS UDP 5353 for local discovery.")
    note("Reserved — do NOT assign these: " + ", ".join(prereqs.RESERVED_IPS) + " (CDM link-local).")

    band("Management-network cabling (each controller → management switch)")
    for item in (
        "Admin port — each controller → management LAN switch (cable ALL controllers; the shared IP fails over).",
        "iLO port — each controller → the same management network (out-of-band management).",
        "CDM Ethernet port — each controller chassis (OCuLink-to-Ethernet dongle) + at least one drive chassis per rack.",
    ):
        ws.append([item, None, None])

    band("Account, network & time")
    for item in (
        "HPE GreenLake account + workspace; Data Services deployed; Administrator role assigned.",
        "Personal API client created (Client ID / Secret / token URL) — enter these on the Initialisation tab.",
        "Array DNS resolves the global names above; system time within 2 minutes of correct.",
        "Onboarding jump box on the same subnet as the array (for the 169.254.x Cloud Connectivity URL).",
    ):
        ws.append([item, None, None])

    band("Reference")
    ws.append(["HPE Alletra Storage MP B10000 — Network requirement details (Planning / install guide, sd00002405)."])

    ws.column_dimensions["A"].width = 64
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 44


def parse_workbook_bytes(
    data: bytes,
    *,
    mode: RunMode = RunMode.FULL_ONBOARDING,
    selected_steps: list[str] | None = None,
    complete: bool = False,
) -> ParsedInitSheet:
    """Read a filled Initialisation_sheet.xlsx back into (GreenLake creds, ArrayWorkItem).

    ``complete=True`` (the upload path, ADR 0005 revision) validates the FULL superset and always
    parses the Provisioning tab — the sheet is filled once, in full, before a mode is chosen. Passing
    a ``mode`` instead scopes validation to that mode's fields (used where a mode is already known).
    """
    workbook = load_workbook(io.BytesIO(data), data_only=True)
    sheet = workbook[SHEET_NAME] if SHEET_NAME in workbook.sheetnames else workbook.active

    values = _read_tab(sheet, _LABEL_TO_KEY)
    required = all_required_keys() if complete else required_keys_for(mode, selected_steps)
    parsed = _build(values, required)

    if complete or _provisioning_selected(mode, selected_steps):
        parsed.provisioning_intent = _parse_provisioning_tab(workbook)
    return parsed


def _read_tab(sheet, label_to_key: dict[str, str]) -> dict[str, str]:
    """Collect Field->Value pairs from a fillable tab, matching labels to stable keys."""
    values: dict[str, str] = {}
    for row in sheet.iter_rows(min_row=1, max_col=2, values_only=True):
        label = row[0]
        value = row[1] if len(row) > 1 else None
        if label is None or value is None:
            continue
        key = label_to_key.get(_normalize_label(str(label)))
        text = str(value).strip()
        if key and text:
            values[key] = text
    return values


def _read_table(ws, columns: list[tuple[str, str, bool]]) -> list[dict[str, str]]:
    """Read a row-table tab into a list of {key: value} dicts. Locates the header row by matching cells
    to the known column headers (tolerant of the intro line and of the operator's column order), then
    reads data rows to the end, skipping any row whose first (name) column is blank."""
    header_to_key = {_normalize_label(label): key for key, label, _ in columns}
    rows = list(ws.iter_rows(values_only=True))
    header_idx: int | None = None
    col_index: dict[int, str] = {}
    for i, row in enumerate(rows):
        mapping = {ci: header_to_key[_normalize_label(str(cell))]
                   for ci, cell in enumerate(row)
                   if cell is not None and _normalize_label(str(cell)) in header_to_key}
        if mapping:
            header_idx, col_index = i, mapping
            break
    if header_idx is None:
        return []

    name_key = columns[0][0]
    records: list[dict[str, str]] = []
    for row in rows[header_idx + 1:]:
        record = {
            key: str(row[ci]).strip()
            for ci, key in col_index.items()
            if ci < len(row) and row[ci] is not None and str(row[ci]).strip()
        }
        if record.get(name_key):
            records.append(record)
    return records


def _parse_provisioning_tab(workbook) -> ProvisioningIntent:
    """Read the provisioning input (Provisioning targets + Volumes + Host sets row-tables) into a plural
    ProvisioningIntent (passwords included — customer-supplied). Exports are composed later in the UI."""
    for name in (PROVISIONING_SHEET_NAME, VOLUMES_SHEET_NAME, HOSTSETS_SHEET_NAME):
        if name not in workbook.sheetnames:
            raise ValueError(f"The '{name}' tab is required for provisioning but was not found in the workbook.")

    targets = _read_tab(workbook[PROVISIONING_SHEET_NAME], _PROV_LABEL_TO_KEY)
    missing = [_PROV_KEY_TO_LABEL[key] for key in _PROV_REQUIRED if not targets.get(key)]
    if missing:
        raise ValueError("Provisioning tab — missing required fields: " + ", ".join(sorted(missing)))

    default_cpg = targets.get("prov_default_cpg") or "SSD_r6"
    default_type = (targets.get("prov_default_type") or "tpvv").strip().lower()
    if default_type not in ("tpvv", "reduce"):
        raise ValueError("Default provisioning type must be 'tpvv' or 'reduce'.")

    vol_rows = _read_table(workbook[VOLUMES_SHEET_NAME], VOLUME_COLUMNS)
    if not vol_rows:
        raise ValueError("Volumes tab — add at least one volume (a name and a size).")
    volumes: list[VolumeRequest] = []
    for r in vol_rows:
        ptype = (r.get("provisioning_type") or default_type).strip().lower()
        if ptype not in ("tpvv", "reduce"):
            raise ValueError(f"Volume '{r['name']}' — Type must be 'tpvv' or 'reduce'.")
        try:
            size_gib = int(float(r["size_gib"]))
        except (KeyError, ValueError) as exc:
            raise ValueError(f"Volume '{r.get('name', '?')}' — Size (GiB) must be a number.") from exc
        volumes.append(VolumeRequest(
            name=r["name"], size_gib=size_gib,
            provisioning_type=ptype,  # type: ignore[arg-type]
            cpg=r.get("cpg") or default_cpg, vvset=r.get("vvset") or None,
        ))

    hs_rows = _read_table(workbook[HOSTSETS_SHEET_NAME], HOSTSET_COLUMNS)
    if not hs_rows:
        raise ValueError("Host sets tab — add at least one host set (a name).")
    host_sets = [
        HostSetRequest(name=r["name"], members=[m.strip() for m in (r.get("members") or "").split(",") if m.strip()])
        for r in hs_rows
    ]

    return ProvisioningIntent(
        array=EndpointCreds(host=targets["prov_array_host"], username=targets["prov_array_user"], password=targets["prov_array_password"]),
        vcenter=EndpointCreds(host=targets["prov_vcenter_host"], username=targets["prov_vcenter_user"], password=targets["prov_vcenter_password"]),
        # Switches are OPTIONAL (verify is array-side; switches only needed to CREATE zones) -> default blank.
        switch_f1=EndpointCreds(host=targets.get("prov_sw1_host", ""), username=targets.get("prov_sw1_user", ""), password=targets.get("prov_sw1_password", "")),
        switch_f2=EndpointCreds(host=targets.get("prov_sw2_host", ""), username=targets.get("prov_sw2_user", ""), password=targets.get("prov_sw2_password", "")),
        volumes=volumes,
        host_sets=host_sets,
    )


def _build(values: dict[str, str], required: set[str]) -> ParsedInitSheet:
    missing = [_KEY_TO_LABEL[key] for key in required if not values.get(key)]
    if missing:
        raise ValueError("Missing required fields: " + ", ".join(sorted(missing)))

    # Fields not required by this mode but still mandatory on the models get harmless placeholders
    # (they are never used by the steps that didn't ask for them).
    def get(key: str, default: str = "") -> str:
        return values.get(key) or default

    dns = [values[key] for key in ("dns1", "dns2", "dns3") if values.get(key)]
    network = NetworkConfig(
        mgmt_ipv4=get("mgmt_ipv4"),
        mask=get("mask"),
        gateway=get("gateway"),
        dns=dns,
        ntp=get("ntp"),
        timezone=get("timezone"),
        proxy_host=values.get("proxy_host") or None,
        proxy_port=int(float(values["proxy_port"])) if values.get("proxy_port") else None,
    )
    dscc_setup = DsccSetupConfig(
        system_name=get("dscc_system_name", get("serial_number", "array")),
        country=get("dscc_country", "—"),
        credential_name=values.get("secret_name") or "b10000-admin",
        username=values.get("secret_username") or "3paradm",
        password=values.get("secret_password") or None,
        contact_first_name=values.get("contact_first_name"),
        contact_last_name=values.get("contact_last_name"),
        contact_language=values.get("contact_language") or "English",
        contact_company=values.get("contact_company"),
        contact_phone=values.get("contact_phone"),
        contact_email=values.get("contact_email"),
    )
    work_item = ArrayWorkItem(
        serial_number=values["serial_number"],  # always required
        part_number=get("part_number", "—"),
        subscription_key=get("subscription_key", "—"),
        service_catalog_region_id=get("service_catalog_region_id", "—"),
        dscc_region_code=get("dscc_region_code", "—"),
        cloudinit_url=values.get("cloudinit_url") or DEFAULT_CLOUDINIT_URL,
        network=network,
        dscc_setup=dscc_setup,
        customer_name=values.get("customer_name", ""),
        site=values.get("site", ""),
    )
    return ParsedInitSheet(
        work_item=work_item,
        gl_client_id=get("gl_client_id"),
        gl_client_secret=get("gl_client_secret"),
        gl_token_url=get("gl_token_url"),
        customer_name=values.get("customer_name", ""),
        site=values.get("site", ""),
    )
